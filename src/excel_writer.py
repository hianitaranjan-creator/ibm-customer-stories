"""
excel_writer.py
---------------
Writes all processed data into a tidy Excel workbook with seven sheets.

Sheet 1 — Executive Summary
Sheet 2 — Story Inventory
Sheet 3 — Proof Inventory
Sheet 4 — Coverage Matrix
Sheet 5 — Evidence Pipeline
Sheet 6 — QA Exceptions
Sheet 7 — Run Log
"""

import os
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

from src.config import GTM_MOTIONS, GEOGRAPHIES, PROOF_STRENGTHS
from src import logger

# ── Colour palette ──────────────────────────────────────────────────────────
_IBM_BLUE   = "054ADA"
_IBM_CYAN   = "1192E8"
_WHITE      = "FFFFFF"
_GREY_LIGHT = "F4F4F4"
_GREY_MID   = "E0E0E0"
_GREEN      = "198038"
_AMBER      = "B28600"
_RED        = "DA1E28"
_AMBER_FILL = "FCF4D6"
_RED_FILL   = "FFF1F1"
_GREEN_FILL = "DEFBE6"


def _header_cell(ws, row, col, text, bg=_IBM_BLUE, fg=_WHITE, bold=True):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font      = Font(bold=bold, color=fg, size=11)
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return cell


def _data_cell(ws, row, col, value, wrap=False, bold=False):
    cell = ws.cell(row=row, column=col, value=value)
    cell.alignment = Alignment(vertical="top", wrap_text=wrap)
    if bold:
        cell.font = Font(bold=True)
    return cell


def _set_col_widths(ws, widths: list[int]):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _freeze(ws, cell="A2"):
    ws.freeze_panes = cell


def _stripe(ws, start_row, end_row, col_count):
    """Alternate light grey rows for readability."""
    for r in range(start_row, end_row + 1):
        if r % 2 == 0:
            for c in range(1, col_count + 1):
                ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=_GREY_LIGHT)


# ─────────────────────────────────────────────────────────────────────────────
# Sheet 1 — Executive Summary
# ─────────────────────────────────────────────────────────────────────────────

def _write_exec_summary(ws, summary: dict):
    ws.title = "Executive Summary"
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, [32, 22])

    # Title banner.
    ws.merge_cells("A1:B1")
    cell = ws["A1"]
    cell.value = "IBM Customer Stories — Executive Summary"
    cell.font  = Font(bold=True, size=14, color=_WHITE)
    cell.fill  = PatternFill("solid", fgColor=_IBM_BLUE)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Sub-title: run timestamp.
    ws.merge_cells("A2:B2")
    ws["A2"].value = f"Last run: {summary.get('run_timestamp', '')}"
    ws["A2"].font  = Font(italic=True, size=10, color="444444")
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 18

    rows = [
        ("", ""),
        ("Stories", ""),
        ("  Total stories",           summary.get("total_stories", 0)),
        ("  Named customers",          summary.get("named_stories", 0)),
        ("  Unnamed customers",        summary.get("unnamed_stories", 0)),
        ("", ""),
        ("Proof Points", ""),
        ("  Total proof points",       summary.get("total_proofs", 0)),
        ("  Strong",                   summary.get("strong_proofs", 0)),
        ("  Medium",                   summary.get("medium_proofs", 0)),
        ("  Weak",                     summary.get("weak_proofs", 0)),
        ("  Restricted",               summary.get("restricted_proofs", 0)),
        ("", ""),
        ("Evidence Gaps", ""),
        ("  Total gaps identified",    summary.get("total_gaps", 0)),
        ("  High-priority gaps",       summary.get("high_priority_gaps", 0)),
        ("", ""),
        ("GTM Motion Coverage (% geos with ≥1 Strong proof)", ""),
    ]

    for motion in GTM_MOTIONS:
        rows.append((f"  {motion}", summary.get("gtm_coverage", {}).get(motion, "0%")))

    rows += [
        ("", ""),
        ("Top 5 Industries by Story Count", ""),
    ]
    for industry, count in summary.get("top5_industries", []):
        rows.append((f"  {industry}", count))

    for i, (label, value) in enumerate(rows, start=3):
        if value == "" and label:
            # Section header.
            cell = ws.cell(row=i, column=1, value=label)
            cell.font = Font(bold=True, size=11, color=_IBM_BLUE)
            ws.merge_cells(f"A{i}:B{i}")
        else:
            ws.cell(row=i, column=1, value=label).alignment = Alignment(indent=1)
            ws.cell(row=i, column=2, value=value)


# ─────────────────────────────────────────────────────────────────────────────
# Sheet 2 — Story Inventory
# ─────────────────────────────────────────────────────────────────────────────

STORY_HEADERS = [
    "Story ID", "Customer Name", "Named / Unnamed", "Client Zero",
    "Business Partner", "Industry", "Sub-Industry", "Geography",
    "IBM Products", "GTM Motions", "Open/Governed/Hybrid",
    "Structured/Unstructured", "Proof Strength", "Publication Date",
    "Age (months)", "Source URL", "QA Flag",
]
STORY_WIDTHS = [10, 24, 16, 12, 17, 22, 18, 18, 30, 42, 22, 22, 15, 16, 14, 50, 18]


def _write_story_inventory(ws, stories: list[dict]):
    ws.title = "Story Inventory"
    _freeze(ws, "A2")
    _set_col_widths(ws, STORY_WIDTHS)

    for c, h in enumerate(STORY_HEADERS, start=1):
        _header_cell(ws, 1, c, h)

    for r, story in enumerate(stories, start=2):
        row = [
            story.get("story_id"),
            story.get("customer_name"),
            story.get("named_unnamed"),
            story.get("client_zero"),
            story.get("business_partner"),
            story.get("industry"),
            story.get("sub_industry"),
            story.get("geography"),
            story.get("products"),
            story.get("gtm_motions"),
            story.get("open_governed_hybrid"),
            story.get("structured_unstructured"),
            story.get("proof_strength"),
            story.get("publication_date"),
            story.get("publication_age_months"),
            story.get("source_url"),
            story.get("qa_flag"),
        ]
        for c, val in enumerate(row, start=1):
            _data_cell(ws, r, c, val, wrap=(c in (9, 10, 16)))

        # Colour-code proof strength.
        ps_cell = ws.cell(row=r, column=13)
        strength = story.get("proof_strength", "")
        _colour_strength_cell(ps_cell, strength)

    _stripe(ws, 2, len(stories) + 1, len(STORY_HEADERS))


# ─────────────────────────────────────────────────────────────────────────────
# Sheet 3 — Proof Inventory
# ─────────────────────────────────────────────────────────────────────────────

PROOF_HEADERS = [
    "Proof ID", "Story ID", "Customer Name", "Proof Type", "Proof Text",
    "Result Type", "Quantified Result", "IBM Solution Credited",
    "GTM Motion(s)", "Proof Strength", "Source URL", "QA Flag",
]
PROOF_WIDTHS = [10, 10, 24, 20, 60, 14, 18, 24, 42, 15, 50, 18]


def _write_proof_inventory(ws, proof_points: list[dict]):
    ws.title = "Proof Inventory"
    _freeze(ws, "A2")
    _set_col_widths(ws, PROOF_WIDTHS)

    for c, h in enumerate(PROOF_HEADERS, start=1):
        _header_cell(ws, 1, c, h)

    for r, proof in enumerate(proof_points, start=2):
        motions = proof.get("gtm_motions", [])
        motions_str = "; ".join(motions) if isinstance(motions, list) else motions

        row = [
            proof.get("proof_id"),
            proof.get("story_id"),
            proof.get("customer_name"),
            proof.get("proof_type"),
            proof.get("proof_text"),
            proof.get("result_type"),
            proof.get("quantified_result"),
            proof.get("ibm_solution_credited"),
            motions_str,
            proof.get("proof_strength"),
            proof.get("source_url"),
            proof.get("qa_flag"),
        ]
        for c, val in enumerate(row, start=1):
            _data_cell(ws, r, c, val, wrap=(c in (5, 9, 11)))

        ps_cell = ws.cell(row=r, column=10)
        _colour_strength_cell(ps_cell, proof.get("proof_strength", ""))

    _stripe(ws, 2, len(proof_points) + 1, len(PROOF_HEADERS))


# ─────────────────────────────────────────────────────────────────────────────
# Sheet 4 — Coverage Matrix
# ─────────────────────────────────────────────────────────────────────────────

def _write_coverage_matrix(ws, matrix: dict):
    ws.title = "Coverage Matrix"
    ws.sheet_view.showGridLines = False

    # Title.
    ws.merge_cells("A1:H1")
    ws["A1"].value = "Coverage Matrix — Stories with proof per GTM motion × Geography"
    ws["A1"].font  = Font(bold=True, size=12, color=_WHITE)
    ws["A1"].fill  = PatternFill("solid", fgColor=_IBM_BLUE)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    # Column headers: geographies.
    _header_cell(ws, 2, 1, "GTM Motion", bg=_IBM_CYAN, fg=_WHITE)
    _header_cell(ws, 2, 2, "Strength",   bg=_IBM_CYAN, fg=_WHITE)
    for c, geo in enumerate(GEOGRAPHIES, start=3):
        _header_cell(ws, 2, c, geo, bg=_IBM_CYAN, fg=_WHITE)

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 13
    for c in range(3, 3 + len(GEOGRAPHIES)):
        ws.column_dimensions[get_column_letter(c)].width = 17

    row = 3
    for motion in GTM_MOTIONS:
        first_in_motion = True
        for strength in PROOF_STRENGTHS:
            if first_in_motion:
                cell = ws.cell(row=row, column=1, value=motion)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(vertical="top")
                first_in_motion = False
            else:
                ws.cell(row=row, column=1, value="")

            ws.cell(row=row, column=2, value=strength)

            for c, geo in enumerate(GEOGRAPHIES, start=3):
                count = matrix.get(motion, {}).get(geo, {}).get(strength, 0)
                cell  = ws.cell(row=row, column=c, value=count)
                cell.alignment = Alignment(horizontal="center")
                # Colour: green ≥2, amber =1, red =0 (only for Strong row).
                if strength == "Strong":
                    if count >= 2:
                        cell.fill = PatternFill("solid", fgColor=_GREEN_FILL)
                        cell.font = Font(color=_GREEN, bold=True)
                    elif count == 1:
                        cell.fill = PatternFill("solid", fgColor=_AMBER_FILL)
                        cell.font = Font(color=_AMBER)
                    else:
                        cell.fill = PatternFill("solid", fgColor=_RED_FILL)
                        cell.font = Font(color=_RED)
            row += 1

        # Blank spacer row between motions.
        row += 1

    # Legend.
    row += 1
    ws.cell(row=row, column=1, value="Colour key (Strong row only):").font = Font(bold=True)
    row += 1
    ws.cell(row=row, column=1, value="  Green = ≥2 Strong proofs")
    ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor=_GREEN_FILL)
    row += 1
    ws.cell(row=row, column=1, value="  Amber = 1 Strong proof")
    ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor=_AMBER_FILL)
    row += 1
    ws.cell(row=row, column=1, value="  Red = 0 Strong proofs")
    ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor=_RED_FILL)


# ─────────────────────────────────────────────────────────────────────────────
# Sheet 5 — Evidence Pipeline
# ─────────────────────────────────────────────────────────────────────────────

PIPELINE_HEADERS = [
    "Gap ID", "Gap Description", "Priority", "GTM Motion", "Geography",
    "Industry", "Strong #", "Medium #", "Weak #",
    "Suggested Story Type", "Upgradeable Stories", "Notes",
]
PIPELINE_WIDTHS = [8, 52, 10, 38, 18, 18, 9, 9, 8, 46, 40, 30]

_PRIORITY_COLOURS = {
    "High":   ("DA1E28", "FFF1F1"),
    "Medium": ("B28600", "FCF4D6"),
    "Low":    ("198038", "DEFBE6"),
}


def _write_evidence_pipeline(ws, pipeline: list[dict]):
    ws.title = "Evidence Pipeline"
    _freeze(ws, "A2")
    _set_col_widths(ws, PIPELINE_WIDTHS)

    for c, h in enumerate(PIPELINE_HEADERS, start=1):
        _header_cell(ws, 1, c, h)

    for r, gap in enumerate(pipeline, start=2):
        row = [
            gap.get("gap_id"),
            gap.get("gap_description"),
            gap.get("priority"),
            gap.get("gtm_motion"),
            gap.get("geography"),
            gap.get("industry"),
            gap.get("strong_count"),
            gap.get("medium_count"),
            gap.get("weak_count"),
            gap.get("suggested_story_type"),
            gap.get("upgradeable_stories"),
            gap.get("notes"),
        ]
        for c, val in enumerate(row, start=1):
            _data_cell(ws, r, c, val, wrap=(c in (2, 10, 11)))

        # Colour the priority cell.
        priority = gap.get("priority", "")
        if priority in _PRIORITY_COLOURS:
            fg, bg = _PRIORITY_COLOURS[priority]
            cell = ws.cell(row=r, column=3)
            cell.font = Font(bold=True, color=fg)
            cell.fill = PatternFill("solid", fgColor=bg)


# ─────────────────────────────────────────────────────────────────────────────
# Sheet 6 — QA Exceptions
# ─────────────────────────────────────────────────────────────────────────────

QA_HEADERS = [
    "Story ID", "Field Name", "Raw Text Found", "Suggested Value", "Reviewer Action",
]
QA_WIDTHS = [10, 22, 60, 30, 30]


def _write_qa_exceptions(ws, qa_list: list[dict]):
    ws.title = "QA Exceptions"
    _freeze(ws, "A2")
    _set_col_widths(ws, QA_WIDTHS)

    for c, h in enumerate(QA_HEADERS, start=1):
        _header_cell(ws, 1, c, h)

    if not qa_list:
        ws.cell(row=2, column=1, value="No QA exceptions recorded.")
        return

    for r, item in enumerate(qa_list, start=2):
        row = [
            item.get("story_id"),
            item.get("field"),
            item.get("raw_text_found"),
            item.get("suggested_value"),
            item.get("reviewer_action"),
        ]
        for c, val in enumerate(row, start=1):
            _data_cell(ws, r, c, val, wrap=(c == 3))

    _stripe(ws, 2, len(qa_list) + 1, len(QA_HEADERS))


# ─────────────────────────────────────────────────────────────────────────────
# Sheet 7 — Run Log
# ─────────────────────────────────────────────────────────────────────────────

RUN_LOG_HEADERS = [
    "Run Timestamp", "Mode", "Pages Attempted", "Pages Succeeded",
    "Pages Failed", "Stories Processed", "Proof Points Extracted",
    "Duration (s)", "Notes",
]
RUN_LOG_WIDTHS = [20, 10, 18, 18, 14, 18, 22, 12, 40]


def _write_run_log(ws, run_stats: dict | None):
    ws.title = "Run Log"
    _freeze(ws, "A2")
    _set_col_widths(ws, RUN_LOG_WIDTHS)

    for c, h in enumerate(RUN_LOG_HEADERS, start=1):
        _header_cell(ws, 1, c, h)

    if not run_stats:
        return

    row = [
        run_stats.get("run_timestamp"),
        run_stats.get("mode"),
        run_stats.get("pages_attempted"),
        run_stats.get("pages_succeeded"),
        run_stats.get("pages_failed"),
        run_stats.get("stories_processed"),
        run_stats.get("proof_points_extracted"),
        run_stats.get("duration_seconds"),
        run_stats.get("notes"),
    ]
    for c, val in enumerate(row, start=1):
        _data_cell(ws, 2, c, val)


# ─────────────────────────────────────────────────────────────────────────────
# Shared helper
# ─────────────────────────────────────────────────────────────────────────────

def _colour_strength_cell(cell, strength: str):
    colours = {
        "Strong":     (_GREEN,  _GREEN_FILL),
        "Medium":     ("0043CE", "EDF5FF"),
        "Weak":       (_AMBER,  _AMBER_FILL),
        "Restricted": ("6F6F6F", _GREY_LIGHT),
    }
    if strength in colours:
        fg, bg = colours[strength]
        cell.font = Font(bold=True, color=fg)
        cell.fill = PatternFill("solid", fgColor=bg)


# ─────────────────────────────────────────────────────────────────────────────
# Main entry point
# ─────────────────────────────────────────────────────────────────────────────

def write_workbook(
    output_path: str,
    stories: list[dict],
    proof_points: list[dict],
    matrix: dict,
    pipeline: list[dict],
    qa_exceptions: list[dict],
    run_stats: dict | None,
    summary: dict,
) -> None:
    """Write all seven sheets to output_path."""
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    wb = openpyxl.Workbook()
    # Remove the default empty sheet.
    default_sheet = wb.active
    wb.remove(default_sheet)

    _write_exec_summary(wb.create_sheet("Executive Summary"), summary)
    _write_story_inventory(wb.create_sheet("Story Inventory"), stories)
    _write_proof_inventory(wb.create_sheet("Proof Inventory"), proof_points)
    _write_coverage_matrix(wb.create_sheet("Coverage Matrix"), matrix)
    _write_evidence_pipeline(wb.create_sheet("Evidence Pipeline"), pipeline)
    _write_qa_exceptions(wb.create_sheet("QA Exceptions"), qa_exceptions)
    _write_run_log(wb.create_sheet("Run Log"), run_stats)

    wb.save(output_path)
    logger.info(f"Excel workbook saved: {output_path}")
